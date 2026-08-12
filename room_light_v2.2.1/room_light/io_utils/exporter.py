"""
io_utils/exporter.py — 分析结果导出（Excel + PNG）
提供 save_dialog() 工具让用户选择路径，也支持直接传路径静默保存。
"""
from __future__ import annotations
import os
import datetime
from typing import Optional

from core.daylight import DaylightResult
from core.models import RoomModel
from io_utils.weather_data import WeatherDataset


def export_excel(
    path: str,
    result: DaylightResult,
    room: RoomModel,
    weather: Optional[WeatherDataset] = None,
) -> str:
    """
    导出 Excel 报告，包含：
      • Sheet1: 汇总指标
      • Sheet2: 网格照度矩阵
      • Sheet3: 网格 DF 矩阵
      • Sheet4: 气象数据（若提供）
    返回路径字符串（成功）或抛出异常。
    """
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    import numpy as np

    wb = openpyxl.Workbook()

    # ── 配色 ──────────────────────────────────────────────────────────────
    _DARK    = "FFffffff"   # white cell bg
    _HEADER  = "FFf5f6f8"   # light gray header
    _ACCENT  = "FF2563eb"   # academic blue
    _GREEN   = "FF16a34a"   # green (pass)
    _RED     = "FFdc2626"   # red (fail)
    _YELLOW  = "FFd97706"   # amber (warning)
    _TXTPRI  = "FF1a1e2e"   # near-black text
    _TXTSEC  = "FF5a6175"   # secondary gray

    def _hfont(bold=True, color=_TXTPRI, size=11):
        return Font(name="Microsoft YaHei", bold=bold,
                    color=color, size=size)
    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)
    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def _border():
        s = Side(style="thin", color="FF353d55")
        return Border(left=s, right=s, top=s, bottom=s)

    def _header_row(ws, row_idx, cols):
        for c, txt in enumerate(cols, 1):
            cell = ws.cell(row_idx, c, txt)
            cell.font      = _hfont(bold=True, color=_TXTPRI, size=10)
            cell.fill      = _fill(_HEADER)
            cell.alignment = _center()
            cell.border    = _border()

    def _data_cell(ws, r, c, val, fmt=None, ok=None):
        cell = ws.cell(r, c, val)
        cell.font      = _hfont(bold=False, size=10)
        cell.fill      = _fill(_DARK)
        cell.alignment = _center()
        cell.border    = _border()
        if ok is True:
            cell.font = _hfont(color=_GREEN, bold=True, size=10)
        elif ok is False:
            cell.font = _hfont(color=_RED, bold=True, size=10)
        if fmt:
            cell.number_format = fmt

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 1: 汇总
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "采光分析汇总"
    ws1.sheet_view.showGridLines = False

    ts  = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws1.cell(1, 1, "建筑室内采光分析报告").font = _hfont(size=14, color=_ACCENT)
    ws1.cell(2, 1, f"生成时间: {ts}").font = _hfont(bold=False, color=_TXTSEC, size=10)
    ws1.merge_cells("A1:D1"); ws1.merge_cells("A2:D2")

    # 房间参数
    row = 4
    ws1.cell(row, 1, "● 房间参数").font = _hfont(color=_ACCENT, size=11)
    row += 1
    room_params = [
        ("长度 L",   f"{room.length/1000:.3f} m"),
        ("宽度 W",   f"{room.width/1000:.3f} m"),
        ("高度 H",   f"{room.height/1000:.3f} m"),
        ("窗户数量", str(len(room.windows))),
        ("墙面反射率 ρw", f"{room.material.rho_wall:.2f}"),
        ("顶棚反射率 ρc", f"{room.material.rho_ceiling:.2f}"),
        ("地面反射率 ρf", f"{room.material.rho_floor:.2f}"),
        ("加权平均反射率 ρ̄", f"{result.rho_bar:.3f}" if result.rho_bar else "—"),
    ]
    _header_row(ws1, row, ["参数", "数值", "", ""])
    row += 1
    for name, val in room_params:
        _data_cell(ws1, row, 1, name)
        _data_cell(ws1, row, 2, val)
        row += 1

    # 窗户明细
    row += 1
    ws1.cell(row, 1, "● 窗户明细").font = _hfont(color=_ACCENT, size=11)
    row += 1
    _header_row(ws1, row,
                ["编号", "朝向", "X(mm)", "Y(mm)", "宽(mm)", "高(mm)", "透射比τ"])
    row += 1
    from core.models import WALL_MAP_R
    for w in room.windows:
        for c, v in enumerate([
            f"W{w.id}", WALL_MAP_R.get(w.wall, w.wall),
            w.x, w.y, w.width, w.height, w.tau
        ], 1):
            _data_cell(ws1, row, c, v)
        row += 1

    # 采光指标
    row += 1
    ws1.cell(row, 1, "● 采光分析结果").font = _hfont(color=_ACCENT, size=11)
    row += 1
    _header_row(ws1, row, ["指标", "计算值", "合格标准", "判定"])
    row += 1
    E_out = result.E_out or 15000
    metrics = [
        ("平均照度 Eavg",   f"{result.E_avg:.1f} lux",  "≥ 300 lux (GB 50033 III类)",
         result.compliant_300),
        ("最低照度 Emin",   f"{result.E_min:.1f} lux",  "—",  None),
        ("最高照度 Emax",   f"{result.E_max:.1f} lux",  "—",  None),
        ("均匀度 U₀",       f"{result.U0:.3f}",          "≥ 0.70",  result.compliant_u0),
        ("平均采光系数 DF_avg", f"{result.DF_avg:.3f}%","≥ 2.0%",  result.DF_avg >= 2.0 if result.DF_avg else None),
        ("最低采光系数 DF_min", f"{result.DF_min:.3f}%","—",  None),
        ("室外照度 Eout",   f"{E_out:.0f} lux",          "CIE 全阴天典型值", None),
        ("计算方法",        result.method or "—",        "—",  None),
    ]
    for name, val, std, ok in metrics:
        _data_cell(ws1, row, 1, name)
        _data_cell(ws1, row, 2, val, ok=ok)
        _data_cell(ws1, row, 3, std)
        verdict = ("✓ 合格" if ok else "✗ 不合格") if ok is not None else "—"
        _data_cell(ws1, row, 4, verdict,
                   ok=(ok if ok is not None else None))
        row += 1

    # Lynes 快速估算对比
    q = result.quick or {}
    if q.get("E_avg"):
        row += 1
        ws1.cell(row, 1, "● Lynes Flux Method 解析对比").font = _hfont(color=_YELLOW, size=11)
        row += 1
        _header_row(ws1, row, ["参数", "数值"])
        row += 1
        for k, v in [
            ("估算平均照度", f"{q['E_avg']:.1f} lux"),
            ("估算 DF_avg",  f"{q['DF_avg']:.2f}%"),
            ("窗地比 WFR",   f"{q['WFR']:.4f}"),
            ("有效透射比",   f"{q.get('tau_eff',0):.3f}"),
        ]:
            _data_cell(ws1, row, 1, k)
            _data_cell(ws1, row, 2, v)
            row += 1

    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 26
    ws1.column_dimensions["D"].width = 12

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 2: 照度矩阵
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("照度矩阵(lux)")
    ws2.sheet_view.showGridLines = False
    import numpy as np
    E_arr = result.E_lux
    xs    = result.grid_x   # mm  (nx,)
    ys    = result.grid_y   # mm  (ny,)

    # Column headers = x coords
    ws2.cell(1, 1, "Y↓ / X→(mm)").font = _hfont(size=9, color=_TXTSEC)
    ws2.cell(1, 1).fill = _fill(_HEADER)
    for c, x in enumerate(xs, 2):
        ws2.cell(1, c, round(x)).font = _hfont(size=8, color=_TXTSEC)
        ws2.cell(1, c).fill = _fill(_HEADER)
        ws2.cell(1, c).alignment = _center()

    from openpyxl.formatting.rule import ColorScaleRule
    max_e = float(E_arr.max()) if E_arr.size else 1

    for r_idx, y in enumerate(ys):
        row_i = r_idx + 2
        ws2.cell(row_i, 1, round(y)).font = _hfont(size=8, color=_TXTSEC)
        ws2.cell(row_i, 1).fill = _fill(_HEADER)
        ws2.cell(row_i, 1).alignment = _center()
        for c_idx, x in enumerate(xs):
            cell = ws2.cell(row_i, c_idx+2, round(float(E_arr[r_idx, c_idx])))
            cell.alignment = _center()
            cell.font = Font(name="Consolas", size=8)

    # Conditional formatting: colour scale
    last_col = get_column_letter(len(xs)+1)
    last_row = len(ys)+1
    ws2.conditional_formatting.add(
        f"B2:{last_col}{last_row}",
        ColorScaleRule(
            start_type="num", start_value=0,   start_color="FFffffff",
            mid_type="num",   mid_value=300,   mid_color="FF86efac",
            end_type="num",   end_value=int(max_e), end_color="FFef4444",
        )
    )
    for col in range(1, len(xs)+2):
        ws2.column_dimensions[get_column_letter(col)].width = 8

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 3: DF 矩阵
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("采光系数DF(%)")
    ws3.sheet_view.showGridLines = False
    ws3.cell(1, 1, "Y↓ / X→(mm)").font = _hfont(size=9, color=_TXTSEC)
    ws3.cell(1, 1).fill = _fill(_HEADER)
    for c, x in enumerate(xs, 2):
        ws3.cell(1, c, round(x)).font = _hfont(size=8, color=_TXTSEC)
        ws3.cell(1, c).fill = _fill(_HEADER)
        ws3.cell(1, c).alignment = _center()
    for r_idx, y in enumerate(ys):
        row_i = r_idx + 2
        ws3.cell(row_i, 1, round(y)).font = _hfont(size=8, color=_TXTSEC)
        ws3.cell(row_i, 1).fill = _fill(_HEADER)
        ws3.cell(row_i, 1).alignment = _center()
        for c_idx in range(len(xs)):
            cell = ws3.cell(row_i, c_idx+2,
                            round(float(result.DF[r_idx, c_idx]), 3))
            cell.alignment = _center()
            cell.font = Font(name="Consolas", size=8)
    for col in range(1, len(xs)+2):
        ws3.column_dimensions[get_column_letter(col)].width = 8

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 4: 气象数据（可选）
    # ══════════════════════════════════════════════════════════════════════
    if weather and weather.is_valid():
        ws4 = wb.create_sheet("气象数据")
        ws4.sheet_view.showGridLines = False
        ws4.cell(1, 1, "气象数据来源: " + weather.source).font = _hfont(color=_ACCENT)
        _header_row(ws4, 2, ["月份", "GHI (W/m²)", "室外照度 (lux)"])
        for i, (m, ghi, lux) in enumerate(weather.summary_rows()):
            _data_cell(ws4, i+3, 1, m)
            _data_cell(ws4, i+3, 2, float(ghi))
            _data_cell(ws4, i+3, 3, float(lux))
        ws4.column_dimensions["A"].width = 10
        ws4.column_dimensions["B"].width = 16
        ws4.column_dimensions["C"].width = 16

    wb.save(path)
    return path


def export_png(path: str, panel) -> str:
    """从 AnalysisPanel 导出热力图 PNG"""
    panel.save_figure(path, dpi=200)
    return path
