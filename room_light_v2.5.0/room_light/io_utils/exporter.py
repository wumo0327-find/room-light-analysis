"""
io_utils/exporter.py — 分析结果导出（Excel + PNG）
提供 save_dialog() 工具让用户选择路径，也支持直接传路径静默保存。
"""
from __future__ import annotations
import os
import datetime
from typing import Optional

from core.daylight import DaylightResult
from core.models import RoomModel
from io_utils.weather_data import WeatherDataset


def export_excel(
    path: str,
    result: DaylightResult,
    room: RoomModel,
    weather: Optional[WeatherDataset] = None,
) -> str:
    """
    导出 Excel 报告，包含：
      • Sheet1: 汇总指标
      • Sheet2: 网格照度矩阵
      • Sheet3: 网格 DF 矩阵
      • Sheet4: 气象数据（若提供）
    返回路径字符串（成功）或抛出异常。
    """
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    import numpy as np

    wb = openpyxl.Workbook()

    # ── 配色 ──────────────────────────────────────────────────────────────
    _DARK    = "FFffffff"   # white cell bg
    _HEADER  = "FFf5f6f8"   # light gray header
    _ACCENT  = "FF2563eb"   # academic blue
    _GREEN   = "FF16a34a"   # green (pass)
    _RED     = "FFdc2626"   # red (fail)
    _YELLOW  = "FFd97706"   # amber (warning)
    _TXTPRI  = "FF1a1e2e"   # near-black text
    _TXTSEC  = "FF5a6175"   # secondary gray

    def _hfont(bold=True, color=_TXTPRI, size=11):
        return Font(name="Microsoft YaHei", bold=bold,
                    color=color, size=size)
    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)
    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def _border():
        s = Side(style="thin", color="FF353d55")
        return Border(left=s, right=s, top=s, bottom=s)

    def _header_row(ws, row_idx, cols):
        for c, txt in enumerate(cols, 1):
            cell = ws.cell(row_idx, c, txt)
            cell.font      = _hfont(bold=True, color=_TXTPRI, size=10)
            cell.fill      = _fill(_HEADER)
            cell.alignment = _center()
            cell.border    = _border()

    def _data_cell(ws, r, c, val, fmt=None, ok=None):
        cell = ws.cell(r, c, val)
        cell.font      = _hfont(bold=False, size=10)
        cell.fill      = _fill(_DARK)
        cell.alignment = _center()
        cell.border    = _border()
        if ok is True:
            cell.font = _hfont(color=_GREEN, bold=True, size=10)
        elif ok is False:
            cell.font = _hfont(color=_RED, bold=True, size=10)
        if fmt:
            cell.number_format = fmt

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 1: 汇总
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "采光分析汇总"
    ws1.sheet_view.showGridLines = False

    ts  = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws1.cell(1, 1, "建筑室内采光分析报告").font = _hfont(size=14, color=_ACCENT)
    ws1.cell(2, 1, f"生成时间: {ts}").font = _hfont(bold=False, color=_TXTSEC, size=10)
    ws1.merge_cells("A1:D1"); ws1.merge_cells("A2:D2")

    # 房间参数
    row = 4
    ws1.cell(row, 1, "● 房间参数").font = _hfont(color=_ACCENT, size=11)
    row += 1
    room_params = [
        ("长度 L",   f"{room.length/1000:.3f} m"),
        ("宽度 W",   f"{room.width/1000:.3f} m"),
        ("高度 H",   f"{room.height/1000:.3f} m"),
        ("窗户数量", str(len(room.windows))),
        ("墙面反射率 ρw", f"{room.material.rho_wall:.2f}"),
        ("顶棚反射率 ρc", f"{room.material.rho_ceiling:.2f}"),
        ("地面反射率 ρf", f"{room.material.rho_floor:.2f}"),
        ("加权平均反射率 ρ̄", f"{result.rho_bar:.3f}" if result.rho_bar else "—"),
    ]
    _header_row(ws1, row, ["参数", "数值", "", ""])
    row += 1
    for name, val in room_params:
        _data_cell(ws1, row, 1, name)
        _data_cell(ws1, row, 2, val)
        row += 1

    # 窗户明细
    row += 1
    ws1.cell(row, 1, "● 窗户明细").font = _hfont(color=_ACCENT, size=11)
    row += 1
    _header_row(ws1, row,
                ["编号", "朝向", "X(mm)", "Y(mm)", "宽(mm)", "高(mm)", "透射比τ"])
    row += 1
    from core.models import WALL_MAP_R
    for w in room.windows:
        for c, v in enumerate([
            f"W{w.id}", WALL_MAP_R.get(w.wall, w.wall),
            w.x, w.y, w.width, w.height, w.tau
        ], 1):
            _data_cell(ws1, row, c, v)
        row += 1

    # 采光指标
    row += 1
    ws1.cell(row, 1, "● 采光分析结果").font = _hfont(color=_ACCENT, size=11)
    row += 1
    _header_row(ws1, row, ["指标", "计算值", "合格标准", "判定"])
    row += 1
    E_out = result.E_out or 15000
    metrics = [
        ("平均照度 Eavg",   f"{result.E_avg:.1f} lux",  "≥ 300 lux (GB 50033 III类)",
         result.compliant_300),
        ("最低照度 Emin",   f"{result.E_min:.1f} lux",  "—",  None),
        ("最高照度 Emax",   f"{result.E_max:.1f} lux",  "—",  None),
        ("均匀度 U₀",       f"{result.U0:.3f}",          "≥ 0.70",  result.compliant_u0),
        ("平均采光系数 DF_avg", f"{result.DF_avg:.3f}%","≥ 2.0%",  result.DF_avg >= 2.0 if result.DF_avg else None),
        ("最低采光系数 DF_min", f"{result.DF_min:.3f}%","—",  None),
        ("室外照度 Eout",   f"{E_out:.0f} lux",          "CIE 全阴天典型值", None),
        ("计算方法",        result.method or "—",        "—",  None),
    ]
    for name, val, std, ok in metrics:
        _data_cell(ws1, row, 1, name)
        _data_cell(ws1, row, 2, val, ok=ok)
        _data_cell(ws1, row, 3, std)
        verdict = ("✓ 合格" if ok else "✗ 不合格") if ok is not None else "—"
        _data_cell(ws1, row, 4, verdict,
                   ok=(ok if ok is not None else None))
        row += 1

    # Lynes 快速估算对比
    q = result.quick or {}
    if q.get("E_avg"):
        row += 1
        ws1.cell(row, 1, "● Lynes Flux Method 解析对比").font = _hfont(color=_YELLOW, size=11)
        row += 1
        _header_row(ws1, row, ["参数", "数值"])
        row += 1
        for k, v in [
            ("估算平均照度", f"{q['E_avg']:.1f} lux"),
            ("估算 DF_avg",  f"{q['DF_avg']:.2f}%"),
            ("窗地比 WFR",   f"{q['WFR']:.4f}"),
            ("有效透射比",   f"{q.get('tau_eff',0):.3f}"),
        ]:
            _data_cell(ws1, row, 1, k)
            _data_cell(ws1, row, 2, v)
            row += 1

    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 26
    ws1.column_dimensions["D"].width = 12

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 2: 照度矩阵
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("照度矩阵(lux)")
    ws2.sheet_view.showGridLines = False
    import numpy as np
    E_arr = result.E_lux
    xs    = result.grid_x   # mm  (nx,)
    ys    = result.grid_y   # mm  (ny,)

    # Column headers = x coords
    ws2.cell(1, 1, "Y↓ / X→(mm)").font = _hfont(size=9, color=_TXTSEC)
    ws2.cell(1, 1).fill = _fill(_HEADER)
    for c, x in enumerate(xs, 2):
        ws2.cell(1, c, round(x)).font = _hfont(size=8, color=_TXTSEC)
        ws2.cell(1, c).fill = _fill(_HEADER)
        ws2.cell(1, c).alignment = _center()

    from openpyxl.formatting.rule import ColorScaleRule
    max_e = float(E_arr.max()) if E_arr.size else 1

    for r_idx, y in enumerate(ys):
        row_i = r_idx + 2
        ws2.cell(row_i, 1, round(y)).font = _hfont(size=8, color=_TXTSEC)
        ws2.cell(row_i, 1).fill = _fill(_HEADER)
        ws2.cell(row_i, 1).alignment = _center()
        for c_idx, x in enumerate(xs):
            cell = ws2.cell(row_i, c_idx+2, round(float(E_arr[r_idx, c_idx])))
            cell.alignment = _center()
            cell.font = Font(name="Consolas", size=8)

    # Conditional formatting: colour scale
    last_col = get_column_letter(len(xs)+1)
    last_row = len(ys)+1
    ws2.conditional_formatting.add(
        f"B2:{last_col}{last_row}",
        ColorScaleRule(
            start_type="num", start_value=0,   start_color="FFffffff",
            mid_type="num",   mid_value=300,   mid_color="FF86efac",
            end_type="num",   end_value=int(max_e), end_color="FFef4444",
        )
    )
    for col in range(1, len(xs)+2):
        ws2.column_dimensions[get_column_letter(col)].width = 8

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 3: DF 矩阵
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("采光系数DF(%)")
    ws3.sheet_view.showGridLines = False
    ws3.cell(1, 1, "Y↓ / X→(mm)").font = _hfont(size=9, color=_TXTSEC)
    ws3.cell(1, 1).fill = _fill(_HEADER)
    for c, x in enumerate(xs, 2):
        ws3.cell(1, c, round(x)).font = _hfont(size=8, color=_TXTSEC)
        ws3.cell(1, c).fill = _fill(_HEADER)
        ws3.cell(1, c).alignment = _center()
    for r_idx, y in enumerate(ys):
        row_i = r_idx + 2
        ws3.cell(row_i, 1, round(y)).font = _hfont(size=8, color=_TXTSEC)
        ws3.cell(row_i, 1).fill = _fill(_HEADER)
        ws3.cell(row_i, 1).alignment = _center()
        for c_idx in range(len(xs)):
            cell = ws3.cell(row_i, c_idx+2,
                            round(float(result.DF[r_idx, c_idx]), 3))
            cell.alignment = _center()
            cell.font = Font(name="Consolas", size=8)
    for col in range(1, len(xs)+2):
        ws3.column_dimensions[get_column_letter(col)].width = 8

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 4: 气象数据（可选）
    # ══════════════════════════════════════════════════════════════════════
    if weather and weather.is_valid():
        ws4 = wb.create_sheet("气象数据")
        ws4.sheet_view.showGridLines = False
        ws4.cell(1, 1, "气象数据来源: " + weather.source).font = _hfont(color=_ACCENT)
        _header_row(ws4, 2, ["月份", "GHI (W/m²)", "室外照度 (lux)"])
        for i, (m, ghi, lux, _tmp) in enumerate(weather.summary_rows()):
            _data_cell(ws4, i+3, 1, m)
            _data_cell(ws4, i+3, 2, float(ghi))
            _data_cell(ws4, i+3, 3, float(lux))
        ws4.column_dimensions["A"].width = 10
        ws4.column_dimensions["B"].width = 16
        ws4.column_dimensions["C"].width = 16

    wb.save(path)
    return path


def export_png(path: str, panel) -> str:
    """从 AnalysisPanel 导出热力图 PNG"""
    panel.save_figure(path, dpi=200)
    return path


def export_excel_v2(
    path: str,
    daylight_result,
    thermal_result,
    room,
    weather=None,
) -> str:
    """
    v2.5.0 综合 Excel 报告（含热环境数据）
    """
    import openpyxl, datetime, numpy as np
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 如果没有热环境结果则回退到旧版
    if thermal_result is None:
        if daylight_result is not None:
            return export_excel(path, daylight_result, room, weather)
        return path

    wb = openpyxl.Workbook()

    _W = "FFffffff"; _H = "FFf5f6f8"; _A = "FF2563eb"
    _G = "FF16a34a"; _R = "FFdc2626"; _T = "FF1a1e2e"; _S = "FF5a6175"

    def _font(bold=True, color=_T, size=10):
        return Font(name="Microsoft YaHei", bold=bold, color=color, size=size)
    def _fill(c): return PatternFill("solid", fgColor=c)
    def _cen(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
    def _bdr():
        s = Side(style="thin", color="FFd0d5e0")
        return Border(left=s,right=s,top=s,bottom=s)

    def _hrow(ws, r, cols):
        for c,txt in enumerate(cols,1):
            cell = ws.cell(r,c,txt)
            cell.font=_font(True,_S,9); cell.fill=_fill(_H)
            cell.alignment=_cen(); cell.border=_bdr()

    def _dcel(ws, r, c, val, ok=None):
        cell = ws.cell(r,c,val)
        cell.font=_font(False,(_G if ok is True else _R if ok is False else _T),10)
        cell.fill=_fill(_W); cell.alignment=_cen(); cell.border=_bdr()

    # ── Sheet1: 汇总 ──────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "综合分析汇总"
    ws1.sheet_view.showGridLines = False
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws1.cell(1,1,"建筑室内光热环境综合分析报告").font = _font(True,_A,14)
    ws1.cell(2,1,f"生成时间: {ts}").font = _font(False,_S,10)
    ws1.merge_cells("A1:F1"); ws1.merge_cells("A2:F2")

    row = 4
    ws1.cell(row,1,"● 采光指标").font = _font(True,_A,11); row+=1
    if daylight_result:
        _hrow(ws1,row,["指标","值","合格线","判定"]); row+=1
        metrics_d = [
            ("平均照度 Eavg",f"{daylight_result.E_avg:.1f} lux","≥300 lux",daylight_result.compliant_300),
            ("均匀度 U₀",f"{daylight_result.U0:.4f}","≥0.70",daylight_result.compliant_u0),
            ("DF_avg",f"{daylight_result.DF_avg:.4f}%","≥2.0%",
             (daylight_result.DF_avg or 0)>=2.0),
        ]
        for name,val,std,ok in metrics_d:
            _dcel(ws1,row,1,name); _dcel(ws1,row,2,val,ok=ok)
            _dcel(ws1,row,3,std); _dcel(ws1,row,4,"✓合格" if ok else "✗不合格",ok=ok)
            row+=1

    row+=1
    ws1.cell(row,1,"● 热环境指标").font = _font(True,_A,11); row+=1
    if thermal_result:
        tr = thermal_result
        _hrow(ws1,row,["指标","值","合格线","判定"]); row+=1
        metrics_t = [
            ("年均自然室温",f"{tr.T_in_annual_avg:.1f}℃","—",None),
            ("舒适月数",str(tr.comfort_months)+"月","≥7月",tr.comfort_months>=7),
            ("超温月数",str(tr.overheat_months)+"月","≤3月",tr.overheat_months<=3),
            ("欠温月数",str(tr.underheat_months)+"月","≤2月",tr.underheat_months<=2),
            ("H_envelope",f"{tr.H_envelope:.1f} W/K","—",None),
            ("SC_effective",f"{tr.SC_effective:.3f}","—",None),
        ]
        for name,val,std,ok in metrics_t:
            _dcel(ws1,row,1,name); _dcel(ws1,row,2,val,ok=ok)
            _dcel(ws1,row,3,std)
            v_txt = ("✓合格" if ok else "✗不合格") if ok is not None else "—"
            _dcel(ws1,row,4,v_txt,ok=ok); row+=1

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 12

    # ── Sheet2: 逐月热环境 ────────────────────────────────────────────────
    if thermal_result:
        ws2 = wb.create_sheet("逐月热环境")
        ws2.sheet_view.showGridLines = False
        _hrow(ws2,1,["月份","室外温度℃","自然室温℃","太阳得热kW",
                      "外墙吸热kW","内热扰kW","状态"])
        tr = thermal_result
        months_zh = ["1月","2月","3月","4月","5月","6月",
                     "7月","8月","9月","10月","11月","12月"]
        for i in range(12):
            t_in = tr.T_in[i]
            if t_in > 26:   status = "过热"
            elif t_in < 18: status = "过冷"
            else:           status = "舒适"
            ok = status=="舒适"
            row2 = i+2
            for c,v in enumerate([
                months_zh[i],
                round(float(tr.T_out[i]),1),
                round(float(t_in),2),
                round(float(tr.Q_solar[i])/1000,3),
                round(float(tr.Q_wall_solar[i])/1000,3),
                round(float(tr.Q_int[i])/1000,3),
                status,
            ],1):
                _dcel(ws2,row2,c,v,ok=(ok if c==7 else None))
        for col in ["A","B","C","D","E","F","G"]:
            ws2.column_dimensions[col].width = 14

    # ── Sheet3: 采光矩阵 ──────────────────────────────────────────────────
    if daylight_result and daylight_result.E_lux is not None:
        ws3 = wb.create_sheet("采光照度矩阵(lux)")
        ws3.sheet_view.showGridLines = False
        xs = daylight_result.grid_x
        ys = daylight_result.grid_y
        E  = daylight_result.E_lux
        ws3.cell(1,1,"Y↓/X→(mm)").font=_font(False,_S,8)
        ws3.cell(1,1).fill=_fill(_H)
        for c,x in enumerate(xs,2):
            ws3.cell(1,c,round(x)).fill=_fill(_H); ws3.cell(1,c).alignment=_cen()
            ws3.cell(1,c).font=_font(False,_S,8)
        for r,y in enumerate(ys):
            ri = r+2
            ws3.cell(ri,1,round(y)).fill=_fill(_H); ws3.cell(ri,1).alignment=_cen()
            ws3.cell(ri,1).font=_font(False,_S,8)
            for c,x in enumerate(xs):
                cell=ws3.cell(ri,c+2,round(float(E[r,c])))
                cell.alignment=_cen(); cell.font=Font(name="Consolas",size=8)
        from openpyxl.formatting.rule import ColorScaleRule
        last_col=get_column_letter(len(xs)+1); last_row=len(ys)+1
        ws3.conditional_formatting.add(
            f"B2:{last_col}{last_row}",
            ColorScaleRule(start_type="num",start_value=0,start_color="FFffffff",
                           mid_type="num",mid_value=300,mid_color="FF86efac",
                           end_type="num",end_value=int(E.max()),end_color="FFef4444"))
        for col in range(1,len(xs)+2):
            ws3.column_dimensions[get_column_letter(col)].width=8

    # ── Sheet4: 气象数据 ──────────────────────────────────────────────────
    if weather and weather.is_valid():
        ws4 = wb.create_sheet("气象数据")
        ws4.sheet_view.showGridLines = False
        ws4.cell(1,1,"气象数据来源: "+weather.source).font=_font(True,_A,11)
        _hrow(ws4,2,["月份","GHI (W/m²)","照度 (lux)","月均温 (℃)"])
        for i,(m,ghi,lux,tmp) in enumerate(weather.summary_rows()):
            _dcel(ws4,i+3,1,m); _dcel(ws4,i+3,2,float(ghi))
            _dcel(ws4,i+3,3,float(lux)); _dcel(ws4,i+3,4,float(tmp))
        for c in ["A","B","C","D"]:
            ws4.column_dimensions[c].width=16

    wb.save(path)
    return path
