"""
io_utils/exporter.py — 分析结果导出（Excel + PNG）
提供 save_dialog() 工具让用户选择路径，也支持直接传路径静默保存。
"""
from __future__ import annotations
import os
import datetime
from typing import Optional

from core.daylight import DaylightResult
from core.models import RoomModel
from core.models import (
    DEFAULT_SUPPORT_COST_PER_M, DEFAULT_INSTALL_COST_PER_WINDOW,
)
from core.complex_models import SpaceModel
from io_utils.weather_data import WeatherDataset


def _safe_color_scale_max(values) -> int:
    """Return a finite positive Excel colour-scale endpoint.

    Analysis grids may legitimately contain NaN outside an irregular room.
    Converting a NaN maximum to ``int`` used to abort the final Excel step
    after the PNG/CSV files had already been written.
    """
    import numpy as np

    array = np.asarray(values, dtype=float)
    finite = array[np.isfinite(array)]
    if not finite.size:
        return 1
    maximum = float(np.max(finite))
    if not np.isfinite(maximum):
        return 1
    return max(1, int(np.ceil(maximum)))


def export_excel(
    path: str,
    result: DaylightResult,
    room: RoomModel | SpaceModel,
    weather: Optional[WeatherDataset] = None,
) -> str:
    """
    导出 Excel 报告，包含：
      • Sheet1: 汇总指标
      • Sheet2: 网格照度矩阵
      • Sheet3: 网格 DF 矩阵
      • Sheet4: 气象数据（若提供）
    返回路径字符串（成功）或抛出异常。
    """
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    import numpy as np

    wb = openpyxl.Workbook()

    # ── 配色 ──────────────────────────────────────────────────────────────
    _DARK    = "FFffffff"   # white cell bg
    _HEADER  = "FFf5f6f8"   # light gray header
    _ACCENT  = "FF2563eb"   # academic blue
    _GREEN   = "FF16a34a"   # green (pass)
    _RED     = "FFdc2626"   # red (fail)
    _YELLOW  = "FFd97706"   # amber (warning)
    _TXTPRI  = "FF1a1e2e"   # near-black text
    _TXTSEC  = "FF5a6175"   # secondary gray

    def _hfont(bold=True, color=_TXTPRI, size=11):
        return Font(name="Microsoft YaHei", bold=bold,
                    color=color, size=size)
    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)
    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def _border():
        s = Side(style="thin", color="FF353d55")
        return Border(left=s, right=s, top=s, bottom=s)

    def _header_row(ws, row_idx, cols):
        for c, txt in enumerate(cols, 1):
            cell = ws.cell(row_idx, c, txt)
            cell.font      = _hfont(bold=True, color=_TXTPRI, size=10)
            cell.fill      = _fill(_HEADER)
            cell.alignment = _center()
            cell.border    = _border()

    def _data_cell(ws, r, c, val, fmt=None, ok=None):
        cell = ws.cell(r, c, val)
        cell.font      = _hfont(bold=False, size=10)
        cell.fill      = _fill(_DARK)
        cell.alignment = _center()
        cell.border    = _border()
        if ok is True:
            cell.font = _hfont(color=_GREEN, bold=True, size=10)
        elif ok is False:
            cell.font = _hfont(color=_RED, bold=True, size=10)
        if fmt:
            cell.number_format = fmt

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 1: 汇总
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "采光分析汇总"
    ws1.sheet_view.showGridLines = False

    ts  = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws1.cell(1, 1, "建筑室内采光分析报告").font = _hfont(size=14, color=_ACCENT)
    ws1.cell(2, 1, f"生成时间: {ts}").font = _hfont(bold=False, color=_TXTSEC, size=10)
    ws1.merge_cells("A1:D1"); ws1.merge_cells("A2:D2")

    # 房间/复杂空间参数
    row = 4
    ws1.cell(row, 1, "● 空间参数").font = _hfont(color=_ACCENT, size=11)
    row += 1
    if isinstance(room, SpaceModel):
        from core.space_geometry import space_floor_area_mm2

        windows = [
            opening
            for wall in room.wall_segments()
            for opening in wall.windows()
        ]
        room_params = [
            ("空间名称", room.name),
            ("建筑面积", f"{space_floor_area_mm2(room) / 1_000_000:.3f} ㎡"),
            ("高度 H", f"{room.height_mm / 1000:.3f} m"),
            ("边界墙段", str(len(room.wall_segments()))),
            ("窗户数量", str(len(windows))),
            ("墙面反射率 ρw", f"{room.material.rho_wall:.2f}"),
            ("顶棚反射率 ρc", f"{room.material.rho_ceiling:.2f}"),
            ("地面反射率 ρf", f"{room.material.rho_floor:.2f}"),
            ("加权平均反射率 ρ̄", f"{result.rho_bar:.3f}" if result.rho_bar else "—"),
        ]
    else:
        room_params = [
            ("长度 L",   f"{room.length/1000:.3f} m"),
            ("宽度 W",   f"{room.width/1000:.3f} m"),
            ("高度 H",   f"{room.height/1000:.3f} m"),
            ("窗户数量", str(len(room.windows))),
            ("墙面反射率 ρw", f"{room.material.rho_wall:.2f}"),
            ("顶棚反射率 ρc", f"{room.material.rho_ceiling:.2f}"),
            ("地面反射率 ρf", f"{room.material.rho_floor:.2f}"),
            ("加权平均反射率 ρ̄", f"{result.rho_bar:.3f}" if result.rho_bar else "—"),
        ]
    _header_row(ws1, row, ["参数", "数值", "", ""])
    row += 1
    for name, val in room_params:
        _data_cell(ws1, row, 1, name)
        _data_cell(ws1, row, 2, val)
        row += 1

    # 窗户明细
    row += 1
    ws1.cell(row, 1, "● 窗户明细").font = _hfont(color=_ACCENT, size=11)
    row += 1
    _header_row(ws1, row,
                ["编号", "朝向", "X(mm)", "Y(mm)", "宽(mm)", "高(mm)", "透射比τ"])
    row += 1
    if isinstance(room, SpaceModel):
        for wall_index, wall in enumerate(room.wall_segments(), 1):
            for opening in wall.windows():
                for c, v in enumerate([
                    opening.name or opening.id,
                    wall.name or f"墙{wall_index}",
                    opening.offset_mm,
                    opening.sill_height_mm,
                    opening.width_mm,
                    opening.height_mm,
                    opening.visible_transmittance,
                ], 1):
                    _data_cell(ws1, row, c, v)
                row += 1
    else:
        from core.models import WALL_MAP_R
        for w in room.windows:
            for c, v in enumerate([
                f"W{w.id}", WALL_MAP_R.get(w.wall, w.wall),
                w.x, w.y, w.width, w.height, w.tau
            ], 1):
                _data_cell(ws1, row, c, v)
            row += 1

    # 采光指标
    row += 1
    ws1.cell(row, 1, "● 采光分析结果").font = _hfont(color=_ACCENT, size=11)
    row += 1
    _header_row(ws1, row, ["指标", "计算值", "合格标准", "判定"])
    row += 1
    E_out = result.E_out or 15000
    metrics = [
        ("平均照度 Eavg",   f"{result.E_avg:.1f} lux",  "≥ 300 lux (GB 50033 III类)",
         result.compliant_300),
        ("最低照度 Emin",   f"{result.E_min:.1f} lux",  "—",  None),
        ("最高照度 Emax",   f"{result.E_max:.1f} lux",  "—",  None),
        ("均匀度 U₀",       f"{result.U0:.3f}",          "≥ 0.70",  result.compliant_u0),
        ("平均采光系数 DF_avg", f"{result.DF_avg:.3f}%","≥ 2.0%",  result.DF_avg >= 2.0 if result.DF_avg else None),
        ("最低采光系数 DF_min", f"{result.DF_min:.3f}%","—",  None),
        ("室外照度 Eout",   f"{E_out:.0f} lux",          "CIE 全阴天典型值", None),
        ("计算方法",        result.method or "—",        "—",  None),
    ]
    for name, val, std, ok in metrics:
        _data_cell(ws1, row, 1, name)
        _data_cell(ws1, row, 2, val, ok=ok)
        _data_cell(ws1, row, 3, std)
        verdict = ("✓ 合格" if ok else "✗ 不合格") if ok is not None else "—"
        _data_cell(ws1, row, 4, verdict,
                   ok=(ok if ok is not None else None))
        row += 1

    # Lynes 快速估算对比
    q = result.quick or {}
    if q.get("E_avg"):
        row += 1
        ws1.cell(row, 1, "● Lynes Flux Method 解析对比").font = _hfont(color=_YELLOW, size=11)
        row += 1
        _header_row(ws1, row, ["参数", "数值"])
        row += 1
        for k, v in [
            ("估算平均照度", f"{q['E_avg']:.1f} lux"),
            ("估算 DF_avg",  f"{q['DF_avg']:.2f}%"),
            ("窗地比 WFR",   f"{q['WFR']:.4f}"),
            ("有效透射比",   f"{q.get('tau_eff',0):.3f}"),
        ]:
            _data_cell(ws1, row, 1, k)
            _data_cell(ws1, row, 2, v)
            row += 1

    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 26
    ws1.column_dimensions["D"].width = 12

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 2: 照度矩阵
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("照度矩阵(lux)")
    ws2.sheet_view.showGridLines = False
    import numpy as np
    E_arr = result.E_lux
    xs    = result.grid_x   # mm  (nx,)
    ys    = result.grid_y   # mm  (ny,)

    # Column headers = x coords
    ws2.cell(1, 1, "Y↓ / X→(mm)").font = _hfont(size=9, color=_TXTSEC)
    ws2.cell(1, 1).fill = _fill(_HEADER)
    for c, x in enumerate(xs, 2):
        ws2.cell(1, c, round(x)).font = _hfont(size=8, color=_TXTSEC)
        ws2.cell(1, c).fill = _fill(_HEADER)
        ws2.cell(1, c).alignment = _center()

    from openpyxl.formatting.rule import ColorScaleRule
    max_e = _safe_color_scale_max(E_arr)

    for r_idx, y in enumerate(ys):
        row_i = r_idx + 2
        ws2.cell(row_i, 1, round(y)).font = _hfont(size=8, color=_TXTSEC)
        ws2.cell(row_i, 1).fill = _fill(_HEADER)
        ws2.cell(row_i, 1).alignment = _center()
        for c_idx, x in enumerate(xs):
            value = float(E_arr[r_idx, c_idx])
            cell = ws2.cell(
                row_i,
                c_idx + 2,
                round(value) if np.isfinite(value) else None,
            )
            cell.alignment = _center()
            cell.font = Font(name="Consolas", size=8)

    # Conditional formatting: colour scale
    last_col = get_column_letter(len(xs)+1)
    last_row = len(ys)+1
    ws2.conditional_formatting.add(
        f"B2:{last_col}{last_row}",
        ColorScaleRule(
            start_type="num", start_value=0,   start_color="FFffffff",
            mid_type="num",   mid_value=300,   mid_color="FF86efac",
            end_type="num",   end_value=max_e, end_color="FFef4444",
        )
    )
    for col in range(1, len(xs)+2):
        ws2.column_dimensions[get_column_letter(col)].width = 8

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 3: DF 矩阵
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("采光系数DF(%)")
    ws3.sheet_view.showGridLines = False
    ws3.cell(1, 1, "Y↓ / X→(mm)").font = _hfont(size=9, color=_TXTSEC)
    ws3.cell(1, 1).fill = _fill(_HEADER)
    for c, x in enumerate(xs, 2):
        ws3.cell(1, c, round(x)).font = _hfont(size=8, color=_TXTSEC)
        ws3.cell(1, c).fill = _fill(_HEADER)
        ws3.cell(1, c).alignment = _center()
    for r_idx, y in enumerate(ys):
        row_i = r_idx + 2
        ws3.cell(row_i, 1, round(y)).font = _hfont(size=8, color=_TXTSEC)
        ws3.cell(row_i, 1).fill = _fill(_HEADER)
        ws3.cell(row_i, 1).alignment = _center()
        for c_idx in range(len(xs)):
            value = float(result.DF[r_idx, c_idx])
            cell = ws3.cell(
                row_i,
                c_idx + 2,
                round(value, 3) if np.isfinite(value) else None,
            )
            cell.alignment = _center()
            cell.font = Font(name="Consolas", size=8)
    for col in range(1, len(xs)+2):
        ws3.column_dimensions[get_column_letter(col)].width = 8

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 4: 气象数据（可选）
    # ══════════════════════════════════════════════════════════════════════
    if weather and weather.is_valid():
        ws4 = wb.create_sheet("气象数据")
        ws4.sheet_view.showGridLines = False
        ws4.cell(1, 1, "气象数据来源: " + weather.source).font = _hfont(color=_ACCENT)
        _header_row(ws4, 2, ["月份", "GHI (W/m²)", "室外照度 (lux)"])
        for i, (m, ghi, lux, _tmp) in enumerate(weather.summary_rows()):
            _data_cell(ws4, i+3, 1, m)
            _data_cell(ws4, i+3, 2, float(ghi))
            _data_cell(ws4, i+3, 3, float(lux))
        ws4.column_dimensions["A"].width = 10
        ws4.column_dimensions["B"].width = 16
        ws4.column_dimensions["C"].width = 16

    wb.save(path)
    return path


def export_png(path: str, panel) -> str:
    """从 AnalysisPanel 导出热力图 PNG"""
    panel.save_figure(path, dpi=200)
    return path


def export_excel_v2(
    path: str,
    daylight_result,
    thermal_result,
    room,
    weather=None,
    experiment_df=None,
    experiment_params=None,
    optimal_daylight_result=None,
    optimal_thermal_result=None,
    optimal_room=None,
    optimal_label: str = "",
) -> str:
    """
    综合 Excel 报告（含当前模型、参数化实验与最优方案数据）。
    """
    import openpyxl, datetime, numpy as np
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    has_experiment = experiment_df is not None and not experiment_df.empty
    has_optimal = optimal_daylight_result is not None or optimal_thermal_result is not None

    # 没有参数化实验扩展数据时保持旧版导出行为。
    if thermal_result is None and not has_experiment and not has_optimal:
        if daylight_result is not None:
            return export_excel(path, daylight_result, room, weather)
        return path

    wb = openpyxl.Workbook()

    _W = "FFffffff"; _H = "FFf5f6f8"; _A = "FF2563eb"
    _G = "FF16a34a"; _R = "FFdc2626"; _T = "FF1a1e2e"; _S = "FF5a6175"

    def _font(bold=True, color=_T, size=10):
        return Font(name="Microsoft YaHei", bold=bold, color=color, size=size)
    def _fill(c): return PatternFill("solid", fgColor=c)
    def _cen(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
    def _bdr():
        s = Side(style="thin", color="FFd0d5e0")
        return Border(left=s,right=s,top=s,bottom=s)

    def _hrow(ws, r, cols):
        for c,txt in enumerate(cols,1):
            cell = ws.cell(r,c,txt)
            cell.font=_font(True,_S,9); cell.fill=_fill(_H)
            cell.alignment=_cen(); cell.border=_bdr()

    def _dcel(ws, r, c, val, ok=None):
        cell = ws.cell(r,c,val)
        cell.font=_font(False,(_G if ok is True else _R if ok is False else _T),10)
        cell.fill=_fill(_W); cell.alignment=_cen(); cell.border=_bdr()

    # ── Sheet1: 汇总 ──────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "综合分析汇总"
    ws1.sheet_view.showGridLines = False
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws1.cell(1,1,"建筑室内光热环境综合分析报告").font = _font(True,_A,14)
    ws1.cell(2,1,f"生成时间: {ts}").font = _font(False,_S,10)
    ws1.merge_cells("A1:F1"); ws1.merge_cells("A2:F2")

    row = 4
    ws1.cell(row,1,"● 采光指标").font = _font(True,_A,11); row+=1
    if daylight_result:
        _hrow(ws1,row,["指标","值","合格线","判定"]); row+=1
        metrics_d = [
            ("平均照度 Eavg",f"{daylight_result.E_avg:.1f} lux","≥300 lux",daylight_result.compliant_300),
            ("均匀度 U₀",f"{daylight_result.U0:.4f}","≥0.70",daylight_result.compliant_u0),
            ("DF_avg",f"{daylight_result.DF_avg:.4f}%","≥2.0%",
             (daylight_result.DF_avg or 0)>=2.0),
        ]
        for name,val,std,ok in metrics_d:
            _dcel(ws1,row,1,name); _dcel(ws1,row,2,val,ok=ok)
            _dcel(ws1,row,3,std); _dcel(ws1,row,4,"✓合格" if ok else "✗不合格",ok=ok)
            row+=1

    row+=1
    ws1.cell(row,1,"● 热环境指标").font = _font(True,_A,11); row+=1
    if thermal_result:
        tr = thermal_result
        _hrow(ws1,row,["指标","值","合格线","判定"]); row+=1
        metrics_t = [
            ("年均自然室温",f"{tr.T_in_annual_avg:.1f}℃","—",None),
            ("热不舒适度",f"{tr.thermal_discomfort:.2f}℃·月","越低越好",None),
            ("过热累积强度",f"{tr.overheat_degree_months:.2f}℃·月","越低越好",None),
            ("欠热累积强度",f"{tr.underheat_degree_months:.2f}℃·月","越低越好",None),
            ("H_envelope",f"{tr.H_envelope:.1f} W/K","—",None),
            ("SC_effective",f"{tr.SC_effective:.3f}","—",None),
        ]
        for name,val,std,ok in metrics_t:
            _dcel(ws1,row,1,name); _dcel(ws1,row,2,val,ok=ok)
            _dcel(ws1,row,3,std)
            v_txt = ("✓合格" if ok else "✗不合格") if ok is not None else "—"
            _dcel(ws1,row,4,v_txt,ok=ok); row+=1

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 12

    # ── Sheet2: 逐月热环境 ────────────────────────────────────────────────
    if thermal_result:
        ws2 = wb.create_sheet("逐月热环境")
        ws2.sheet_view.showGridLines = False
        _hrow(ws2,1,["月份","室外温度℃","自然室温℃","太阳得热kW",
                      "外墙吸热kW","内热扰kW","状态"])
        tr = thermal_result
        months_zh = ["1月","2月","3月","4月","5月","6月",
                     "7月","8月","9月","10月","11月","12月"]
        for i in range(12):
            t_in = tr.T_in[i]
            if t_in > 26:   status = "过热"
            elif t_in < 18: status = "过冷"
            else:           status = "舒适"
            ok = status=="舒适"
            row2 = i+2
            for c,v in enumerate([
                months_zh[i],
                round(float(tr.T_out[i]),1),
                round(float(t_in),2),
                round(float(tr.Q_solar[i])/1000,3),
                round(float(tr.Q_wall_solar[i])/1000,3),
                round(float(tr.Q_int[i])/1000,3),
                status,
            ],1):
                _dcel(ws2,row2,c,v,ok=(ok if c==7 else None))
        for col in ["A","B","C","D","E","F","G"]:
            ws2.column_dimensions[col].width = 14

    # ── Sheet3: 采光矩阵 ──────────────────────────────────────────────────
    if daylight_result and daylight_result.E_lux is not None:
        ws3 = wb.create_sheet("采光照度矩阵(lux)")
        ws3.sheet_view.showGridLines = False
        xs = daylight_result.grid_x
        ys = daylight_result.grid_y
        E  = daylight_result.E_lux
        ws3.cell(1,1,"Y↓/X→(mm)").font=_font(False,_S,8)
        ws3.cell(1,1).fill=_fill(_H)
        for c,x in enumerate(xs,2):
            ws3.cell(1,c,round(x)).fill=_fill(_H); ws3.cell(1,c).alignment=_cen()
            ws3.cell(1,c).font=_font(False,_S,8)
        for r,y in enumerate(ys):
            ri = r+2
            ws3.cell(ri,1,round(y)).fill=_fill(_H); ws3.cell(ri,1).alignment=_cen()
            ws3.cell(ri,1).font=_font(False,_S,8)
            for c,x in enumerate(xs):
                value = float(E[r, c])
                cell=ws3.cell(
                    ri,
                    c + 2,
                    round(value) if np.isfinite(value) else None,
                )
                cell.alignment=_cen(); cell.font=Font(name="Consolas",size=8)
        from openpyxl.formatting.rule import ColorScaleRule
        last_col=get_column_letter(len(xs)+1); last_row=len(ys)+1
        ws3.conditional_formatting.add(
            f"B2:{last_col}{last_row}",
            ColorScaleRule(start_type="num",start_value=0,start_color="FFffffff",
                           mid_type="num",mid_value=300,mid_color="FF86efac",
                           end_type="num",
                           end_value=_safe_color_scale_max(E),
                           end_color="FFef4444"))
        for col in range(1,len(xs)+2):
            ws3.column_dimensions[get_column_letter(col)].width=8

    # ── Sheet4: 气象数据 ──────────────────────────────────────────────────
    if weather and weather.is_valid():
        ws4 = wb.create_sheet("气象数据")
        ws4.sheet_view.showGridLines = False
        ws4.cell(1,1,"气象数据来源: "+weather.source).font=_font(True,_A,11)
        _hrow(ws4,2,["月份","GHI (W/m²)","照度 (lux)","月均温 (℃)"])
        for i,(m,ghi,lux,tmp) in enumerate(weather.summary_rows()):
            _dcel(ws4,i+3,1,m); _dcel(ws4,i+3,2,float(ghi))
            _dcel(ws4,i+3,3,float(lux)); _dcel(ws4,i+3,4,float(tmp))
        for c in ["A","B","C","D"]:
            ws4.column_dimensions[c].width=16

    # ── 参数化实验：完整算例表 ────────────────────────────────────────────
    if has_experiment:
        import math

        def _excel_value(value):
            """把numpy/pandas标量转换为openpyxl可安全写入的值。"""
            if value is None:
                return None
            if isinstance(value, np.generic):
                value = value.item()
            if isinstance(value, float) and not math.isfinite(value):
                return None
            if isinstance(value, (list, tuple, dict, set)):
                return str(value)
            return value

        exp = experiment_df.copy()
        internal_cols = [c for c in exp.columns if str(c).startswith("_")]
        if internal_cols:
            exp = exp.drop(columns=internal_cols)
        ws5 = wb.create_sheet("参数化实验结果")
        ws5.sheet_view.showGridLines = False
        headers = [str(c) for c in exp.columns]
        _hrow(ws5, 1, headers)
        for r_idx, values in enumerate(exp.itertuples(index=False, name=None), 2):
            for c_idx, value in enumerate(values, 1):
                _dcel(ws5, r_idx, c_idx, _excel_value(value))
        ws5.freeze_panes = "A2"
        if headers:
            ws5.auto_filter.ref = (
                f"A1:{get_column_letter(len(headers))}{len(exp) + 1}")
        for c_idx, header in enumerate(headers, 1):
            sample = [str(header)]
            sample.extend(
                str(_excel_value(v) or "")
                for v in exp.iloc[:100, c_idx - 1].tolist())
            width = min(42, max(10, max(len(v) for v in sample) + 2))
            ws5.column_dimensions[get_column_letter(c_idx)].width = width

        ws6 = wb.create_sheet("参数化实验说明")
        ws6.sheet_view.showGridLines = False
        ws6.cell(1, 1, "参数化实验与造价说明").font = _font(True, _A, 14)
        is_complex_experiment = bool(
            experiment_params
            and experiment_params.get("model_type") == "complex_space"
        )
        engineering_link = (
            "实验直接采用当前.rlproj活动空间的真实多边形边界、墙段朝向、"
            "外窗尺寸/位置、逐窗玻璃光学与热工参数及当前气候；"
            "每组水平遮阳统一应用到该空间全部外窗。"
            if is_complex_experiment else
            "实验采用当前.rlproj的房间尺寸、窗户尺寸/位置、玻璃光学与热工参数，"
            "以及当前气候数据。"
        )
        info_rows = [
            ("工程关联", engineering_link),
            ("改造前基准",
             "完整保留导入.rlproj当前状态计算一次，"
             "用于改造前后对比；新增改造造价记为0元。"),
            ("L=0规则",
             "新方案先移除原遮阳；L=0表示无新增水平遮阳，不随倾角θ、间隙h"
             "或材料重复计算。若原工程本来无遮阳，同一个基准点兼任L=0候选。"),
            ("造价口径",
             "板材面积=当前空间全部外窗宽度合计×板长L；"
             "支撑长度=2×全部外窗宽度合计；总价=板材面积×材料综合单价"
             "+支撑长度×支撑单价+外窗数×安装费。"),
            ("造价性质",
             "软件内置单价是可编辑初始估算值；正式工程造价应以项目所在地、"
             "设计做法、计价期信息价和施工报价复核。"),
            ("湖南计价参考",
             "湖南省建设工程计价办法及消耗量标准（2020）与湖南省建设工程工程量清单"
             "计价标准（2025）；本表不将参考单价冒充最终招标/结算价。"),
            ("2020参考网址",
             "https://zjt.hunan.gov.cn/zjt/hnweb/zxfb/202005/t20200509_12058798.html"),
            ("2025参考网址",
             "https://zjt.hunan.gov.cn/zjt/hnweb/zxfb/202509/t20250926_33815179.html"),
        ]
        if experiment_params:
            info_rows.extend([
                ("材料板综合单价(元/㎡)",
                 str(experiment_params.get("material_unit_costs", {}))),
                ("支撑综合单价(元/m)",
                 float(experiment_params.get(
                     "support_cost_per_m", DEFAULT_SUPPORT_COST_PER_M))),
                ("单窗安装费(元/窗)",
                 float(experiment_params.get(
                     "install_cost_per_window",
                     DEFAULT_INSTALL_COST_PER_WINDOW))),
                ("U₀筛选下限",
                 float(experiment_params.get("u0_min", 0.0))),
            ])
        _hrow(ws6, 3, ["项目", "说明/数值"])
        for r_idx, (name, value) in enumerate(info_rows, 4):
            _dcel(ws6, r_idx, 1, name)
            _dcel(ws6, r_idx, 2, value)
            ws6.cell(r_idx, 2).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True)
        ws6.column_dimensions["A"].width = 24
        ws6.column_dimensions["B"].width = 88

    # ── 最优遮阳方案：汇总、完整逐月热环境与采光矩阵 ─────────────────────
    if has_optimal:
        ws7 = wb.create_sheet("最优方案汇总")
        ws7.sheet_view.showGridLines = False
        ws7.cell(1, 1, "参数化实验决策推荐方案").font = _font(True, _A, 14)
        ws7.cell(2, 1, optimal_label or "决策推荐方案").font = _font(True, _S, 10)
        _hrow(ws7, 4, ["项目", "数值"])
        summary_rows = []
        if has_experiment:
            if "decision_recommended" in experiment_df.columns:
                selected = experiment_df[
                    experiment_df["decision_recommended"].fillna(False).astype(bool)
                ]
            elif "balanced_recommended" in experiment_df.columns:
                selected = experiment_df[
                    experiment_df["balanced_recommended"].fillna(False).astype(bool)
                ]
            else:
                selected = experiment_df.iloc[0:0]
            if not selected.empty:
                selected_row = selected.iloc[0]
                for key, label in [
                    ("material", "材料"),
                    ("tilt_deg", "倾角θ(°)"),
                    ("L_mm", "板长L(mm)"),
                    ("gap_mm", "间隙h(mm)"),
                    ("Ra", "采光达标面积比Ra"),
                    ("daylight_score", "连续采光达标度Cd"),
                    ("U0", "采光均匀度U₀"),
                    ("thermal_discomfort", "热不舒适度Σ(℃·月)"),
                    ("building_type", "建筑类型"),
                    ("room_use", "房间用途"),
                    ("teaching_days_per_year", "年教学日(d)"),
                    ("lighting_hours_per_day", "日照明使用时长(h)"),
                    ("hvac_hours_per_day", "日空调使用时长(h)"),
                    ("annual_lighting_kwh", "年照明用电量(kWh)"),
                    ("annual_cooling_kwh", "年制冷用电量(kWh)"),
                    ("annual_heating_kwh", "年制热用电量(kWh)"),
                    ("annual_hvac_kwh", "年空调用电量(kWh)"),
                    ("baseline_annual_operating_cost", "改造前基准年运行电费(元/年)"),
                    ("annual_operating_cost", "年运行电费(元/年)"),
                    ("annual_operating_saving", "年运行费用节省额(元/年)"),
                    ("operating_saving_rate", "年运行费用节省率"),
                    ("simple_payback_years", "静态投资回收期(年)"),
                    ("construction_cost", "遮阳工程造价(元)"),
                    ("annualized_construction_cost", "遮阳年化造价(元/年)"),
                    ("annual_total_cost", "年综合费用(元/年)"),
                    ("panel_area_m2", "遮阳板面积(㎡)"),
                    ("material_unit_price", "材料综合单价(元/㎡)"),
                    ("material_cost", "材料费(元)"),
                    ("support_length_m", "支撑长度(m)"),
                    ("support_unit_price", "支撑综合单价(元/m)"),
                    ("support_cost", "支撑费(元)"),
                    ("window_count", "计价窗数"),
                    ("install_unit_price", "单窗安装费(元/窗)"),
                    ("installation_cost", "安装费(元)"),
                    ("recommendation_pool", "最终推荐池"),
                    ("decision_kind", "推荐类型"),
                    ("decision_reason", "推荐理由"),
                ]:
                    if key in selected_row.index:
                        summary_rows.append((label, selected_row[key]))
        if optimal_room is not None:
            if isinstance(optimal_room, SpaceModel):
                from core.space_geometry import space_floor_area_mm2

                optimal_windows = [
                    opening
                    for wall in optimal_room.wall_segments()
                    if wall.boundary_type in {"exterior", "ground"}
                    for opening in wall.windows()
                ]
                summary_rows.extend([
                    ("空间名称", optimal_room.name),
                    ("空间面积(㎡)",
                     space_floor_area_mm2(optimal_room) / 1_000_000.0),
                    ("空间高度(mm)", optimal_room.height_mm),
                    ("外墙段数量", len([
                        wall for wall in optimal_room.wall_segments()
                        if wall.boundary_type in {"exterior", "ground"}
                    ])),
                    ("外窗数量", len(optimal_windows)),
                    ("外窗总面积(㎡)",
                     sum(opening.area_m2 for opening in optimal_windows)),
                ])
            else:
                summary_rows.extend([
                    ("房间宽度(mm)", optimal_room.width),
                    ("房间进深(mm)", optimal_room.length),
                    ("房间高度(mm)", optimal_room.height),
                    ("窗户数量", len(optimal_room.windows)),
                    ("窗户总面积(㎡)", optimal_room.total_window_area_m2),
                ])
        if optimal_daylight_result is not None:
            summary_rows.extend([
                ("最优方案平均照度(lux)", optimal_daylight_result.E_avg),
                ("最优方案DF_avg(%)", optimal_daylight_result.DF_avg),
                ("最优方案Ra", optimal_daylight_result.Ra),
                ("最优方案连续采光达标度Cd",
                 getattr(optimal_daylight_result, "daylight_score", None)),
                ("最优方案U₀", optimal_daylight_result.U0),
            ])
        if optimal_thermal_result is not None:
            summary_rows.extend([
                ("最优方案年均自然室温(℃)",
                 optimal_thermal_result.T_in_annual_avg),
                ("最优方案热不舒适度(℃·月)", optimal_thermal_result.thermal_discomfort),
                ("最优方案过热累积强度(℃·月)", optimal_thermal_result.overheat_degree_months),
                ("最优方案欠热累积强度(℃·月)", optimal_thermal_result.underheat_degree_months),
                ("最优方案SC_effective",
                 optimal_thermal_result.SC_effective),
            ])
        for r_idx, (name, value) in enumerate(summary_rows, 5):
            _dcel(ws7, r_idx, 1, name)
            _dcel(ws7, r_idx, 2, _excel_value(value) if has_experiment else value)
        ws7.column_dimensions["A"].width = 36
        ws7.column_dimensions["B"].width = 42

        if optimal_thermal_result is not None:
            ws8 = wb.create_sheet("最优方案逐月热环境")
            ws8.sheet_view.showGridLines = False
            _hrow(ws8, 1, [
                "月份", "室外温度℃", "自然室温℃", "太阳得热kW",
                "外墙吸热kW", "内热扰kW", "状态"])
            tr = optimal_thermal_result
            months_zh = [
                "1月", "2月", "3月", "4月", "5月", "6月",
                "7月", "8月", "9月", "10月", "11月", "12月"]
            for i in range(12):
                t_in = float(tr.T_in[i])
                status = "过热" if t_in > 26 else "过冷" if t_in < 18 else "舒适"
                values = [
                    months_zh[i], float(tr.T_out[i]), t_in,
                    float(tr.Q_solar[i]) / 1000,
                    float(tr.Q_wall_solar[i]) / 1000,
                    float(tr.Q_int[i]) / 1000, status]
                for c_idx, value in enumerate(values, 1):
                    _dcel(
                        ws8, i + 2, c_idx, value,
                        ok=((status == "舒适") if c_idx == 7 else None))
            for col in ["A", "B", "C", "D", "E", "F", "G"]:
                ws8.column_dimensions[col].width = 16

        if (optimal_daylight_result is not None
                and optimal_daylight_result.E_lux is not None):
            ws9 = wb.create_sheet("最优方案采光矩阵(lux)")
            ws9.sheet_view.showGridLines = False
            xs = optimal_daylight_result.grid_x
            ys = optimal_daylight_result.grid_y
            E = optimal_daylight_result.E_lux
            ws9.cell(1, 1, "Y↓/X→(mm)").font = _font(False, _S, 8)
            ws9.cell(1, 1).fill = _fill(_H)
            for c_idx, x in enumerate(xs, 2):
                _dcel(ws9, 1, c_idx, round(float(x)))
            for r_idx, y in enumerate(ys, 2):
                _dcel(ws9, r_idx, 1, round(float(y)))
                for c_idx, value in enumerate(E[r_idx - 2], 2):
                    numeric = float(value)
                    _dcel(
                        ws9, r_idx, c_idx,
                        round(numeric) if np.isfinite(numeric) else None,
                    )
            for col in range(1, len(xs) + 2):
                ws9.column_dimensions[get_column_letter(col)].width = 8

    wb.save(path)
    return path
